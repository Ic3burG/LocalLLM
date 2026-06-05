import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import numpy as np


def _make_docx_bytes(text: str) -> bytes:
    from docx import Document

    doc = Document()
    doc.add_paragraph(text)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _make_xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 95
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_document_returns_chunks_without_embeddings():
    from ingest import parse_document

    result = parse_document(_make_docx_bytes("Hello parallel world."), "a.docx")

    assert result is not None
    assert result["filename"] == "a.docx"
    assert result["chunks"]
    assert "Hello parallel world." in result["chunks"][0]["text"]
    # The parse phase must NOT embed — embedding is a separate batched phase,
    # so no per-worker model load happens inside the process pool.
    assert "embeddings" not in result


def test_parse_document_unknown_type_returns_none():
    from ingest import parse_document

    assert parse_document(b"garbage", "notes.txt") is None


def test_parse_documents_parallel_preserves_order():
    from ingest import parse_documents_parallel

    files = [
        (_make_docx_bytes("alpha content here"), "alpha.docx"),
        (_make_docx_bytes("bravo content here"), "bravo.docx"),
        (_make_docx_bytes("charlie content here"), "charlie.docx"),
    ]

    results = parse_documents_parallel(files, max_workers=2)

    # Order must match the input regardless of which worker finished first.
    assert [r["filename"] for r in results] == [
        "alpha.docx",
        "bravo.docx",
        "charlie.docx",
    ]
    assert "alpha" in results[0]["chunks"][0]["text"]
    assert "charlie" in results[2]["chunks"][0]["text"]


def test_parse_documents_parallel_empty_input():
    from ingest import parse_documents_parallel

    assert parse_documents_parallel([]) == []


# --- Batched-embed phase (ingest_documents) ---------------------------------
#
# embed_texts loads the sentence-transformers model, which is exactly the cost
# we are batching to avoid. We monkeypatch it with a stand-in that records how
# many times it was called and returns row-index-encoded vectors, so the tests
# verify the orchestration's real behaviour: ONE batched call, and embeddings
# sliced back to the right document in input order.


def _counting_embed(counter):
    def fake_embed_texts(texts):
        counter["calls"] += 1
        counter["sizes"].append(len(texts))
        arr = np.array([[float(i)] for i in range(len(texts))], dtype=np.float32)
        return arr, 1.23

    return fake_embed_texts


def test_ingest_documents_uses_single_batched_embed_call(monkeypatch):
    import pdf_pipeline
    from ingest import ingest_documents

    counter = {"calls": 0, "sizes": []}
    monkeypatch.setattr(pdf_pipeline, "embed_texts", _counting_embed(counter))

    files = [
        (_make_docx_bytes("first document body"), "one.docx"),
        (_make_docx_bytes("second document body"), "two.docx"),
    ]
    ingest_documents(files, max_workers=2)

    # All chunks across both docs embed in exactly one call — not one per doc.
    assert counter["calls"] == 1


def test_ingest_documents_slices_embeddings_per_document(monkeypatch):
    import pdf_pipeline
    from ingest import ingest_documents

    counter = {"calls": 0, "sizes": []}
    monkeypatch.setattr(pdf_pipeline, "embed_texts", _counting_embed(counter))

    long_text = " ".join(f"w{i}" for i in range(350))  # multi-chunk body
    files = [
        (_make_docx_bytes(long_text), "long.docx"),
        (_make_docx_bytes("tiny"), "tiny.docx"),
    ]
    results = ingest_documents(files, max_workers=2)

    assert [r["filename"] for r in results] == ["long.docx", "tiny.docx"]
    assert results[0]["doc_id"] != results[1]["doc_id"]

    # Row-index-encoded vectors prove each doc got its own contiguous slice of
    # the single batched embed, in input order — whatever each doc's chunk count
    # turns out to be. At least one doc spans multiple rows.
    assert len(results[0]["embeddings"]) >= 2
    cursor = 0
    for r in results:
        n = len(r["chunks"])
        expected = np.array(
            [[float(i)] for i in range(cursor, cursor + n)], dtype=np.float32
        )
        np.testing.assert_array_equal(r["embeddings"], expected)
        cursor += n


def test_ingest_documents_keeps_none_for_failed_parse(monkeypatch):
    import pdf_pipeline
    from ingest import ingest_documents

    counter = {"calls": 0, "sizes": []}
    monkeypatch.setattr(pdf_pipeline, "embed_texts", _counting_embed(counter))

    files = [
        (_make_docx_bytes("good doc"), "good.docx"),
        (b"not a real document", "notes.txt"),  # unparseable -> None
    ]
    results = ingest_documents(files, max_workers=2)

    assert results[1] is None
    assert results[0]["filename"] == "good.docx"
    assert "embeddings" in results[0]
    # Only the good doc's chunk(s) were embedded, in a single call.
    assert counter["calls"] == 1
    assert counter["sizes"] == [len(results[0]["chunks"])]


def test_ingest_documents_handles_mixed_pdf_office_batch(monkeypatch):
    # The frontend routes .docx and .xlsx through the same batch path as PDFs,
    # so a mixed batch must index every file type via one embed call.
    import pdf_pipeline
    from ingest import ingest_documents

    counter = {"calls": 0, "sizes": []}
    monkeypatch.setattr(pdf_pipeline, "embed_texts", _counting_embed(counter))

    files = [
        (_make_docx_bytes("word document body"), "report.docx"),
        (_make_xlsx_bytes(), "data.xlsx"),
    ]
    results = ingest_documents(files, max_workers=2)

    assert [r["filename"] for r in results] == ["report.docx", "data.xlsx"]
    for r in results:
        assert r["doc_id"]
        assert len(r["embeddings"]) == len(r["chunks"]) >= 1
    # Every file's chunks embed in a single shared call.
    assert counter["calls"] == 1
